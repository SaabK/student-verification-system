import cv2
from pyzbar.pyzbar import decode
import numpy as np
from openpyxl import load_workbook
import face_recognition
import os
import time

def load_cms_ids(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    cms_ids = []
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            cms_ids.append(str(row[0]).strip())
    return cms_ids

def load_face_database(folder_path):
    face_db = {}
    for filename in os.listdir(folder_path):
        if filename.lower().endswith((".jpg", ".png", ".jpeg")):
            cms_id = os.path.splitext(filename)[0]
            image_path = os.path.join(folder_path, filename)
            try:
                bgr_image = cv2.imread(image_path)
                if bgr_image is None:
                    continue
                rgb_image = cv2.cvtColor(bgr_image, cv2.COLOR_BGR2RGB)
                encodings = face_recognition.face_encodings(rgb_image)
                if encodings:
                    face_db[cms_id] = encodings[0]
                    print(f"✅ Loaded: {cms_id}")
            except Exception as e:
                print(f"❌ Error: {filename}")
    return face_db

# DATA ANALYSIS: Track verification times using numpy
verification_times = []

def analyze_performance():
    """Analyze system performance using numpy"""
    if len(verification_times) > 0:
        times_array = np.array(verification_times)
        print(f"\n{'='*40}")
        print("PERFORMANCE ANALYSIS")
        print(f"{'='*40}")
        print(f"Total entries: {len(times_array)}")
        print(f"Average time: {np.mean(times_array):.2f}s")
        print(f"Fastest: {np.min(times_array):.2f}s")
        print(f"Slowest: {np.max(times_array):.2f}s")
        print(f"Std deviation: {np.std(times_array):.2f}s")
        print(f"{'='*40}\n")

face_database = load_face_database("faces")
cap = cv2.VideoCapture(0)
valid_cms_ids = load_cms_ids("./students.xlsx")

waiting_for_face = False
scanned_id = None
verification_start_time = None
VERIFICATION_TIMEOUT = 10

while True:
    ret, frame = cap.read()
    
    if not waiting_for_face:
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        barcodes = decode(gray)

        for barcode in barcodes:
            barcode_data = barcode.data.decode("utf-8")
            if barcode_data in valid_cms_ids:
                scanned_id = barcode_data
                waiting_for_face = True
                verification_start_time = time.time()
                x, y, w, h = barcode.rect
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, "SHOW FACE", (x, y - 10), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            else:
                x, y, w, h = barcode.rect
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.putText(frame, "DENIED", (x, y - 10), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
    else:
        elapsed = time.time() - verification_start_time
        remaining = int(VERIFICATION_TIMEOUT - elapsed)
        
        if elapsed > VERIFICATION_TIMEOUT:
            waiting_for_face = False
            scanned_id = None
            continue
        
        cv2.putText(frame, f"ID: {scanned_id} | Time: {remaining}s", (10, 30), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 0), 2)
        
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        live_faces = face_recognition.face_encodings(rgb_frame)
        
        if live_faces:
            stored_encoding = face_database.get(scanned_id)
            if stored_encoding is not None:
                match = face_recognition.compare_faces([stored_encoding], live_faces[0])[0]
                
                # Record verification time for data analysis
                verify_time = time.time() - verification_start_time
                verification_times.append(verify_time)
                
                if match:
                    result_msg = "ACCESS GRANTED"
                    result_color = (0, 255, 0)
                    print(f"✅ {scanned_id} - ACCESS GRANTED")
                else:
                    result_msg = "FACE MISMATCH"
                    result_color = (0, 0, 255)
                    print(f"❌ {scanned_id} - FACE MISMATCH")
                
                # Show result briefly without blocking
                cv2.putText(frame, result_msg, (10, 90), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1.2, result_color, 3)
                cv2.imshow("Scanner", frame)
                cv2.waitKey(1500)  # Brief 1.5 second display
                
                waiting_for_face = False
                scanned_id = None

    cv2.imshow("Scanner", frame)
    
    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break
    elif key == ord('a'):  # Press 'a' for analysis
        analyze_performance()

cap.release()
cv2.destroyAllWindows()
analyze_performance()  # Show final stats