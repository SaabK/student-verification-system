import cv2 # Used for video and image analysis
from pyzbar.pyzbar import decode # Decode barcodes
from openpyxl import load_workbook
import face_recognition
import os # interact with os
import time # just used for timer

def load_cms_ids(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active # select default sheet 1

    cms_ids = []

    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            cms_ids.append(str(row[0]))

    return cms_ids

def load_face_database(folder_path):
    face_db = {}

    for filename in os.listdir(folder_path):
        if filename.endswith((".jpg", ".png", ".jpeg")):

            cms_id = os.path.splitext(filename)[0]
            image_path = os.path.join(folder_path, filename)

            try:
                bgr_image = cv2.imread(image_path)

                if bgr_image is None:
                    print(f"Could not read image: {filename}") # if image is empty, then move to then next image
                    continue

                rgb_image = cv2.cvtColor(bgr_image, cv2.COLOR_BGR2RGB)
                encodings = face_recognition.face_encodings(rgb_image)

                if encodings:
                    face_db[cms_id] = encodings[0]
                    print(f"Loaded face for CMS ID: {cms_id}")
                else:
                    print(f"No face found in image: {filename}")

            except Exception as e:
                print(f"Error processing {filename}: {e}")

    return face_db


face_database = load_face_database("faces")
cap = cv2.VideoCapture(1) # Display video capture
valid_cms_ids = load_cms_ids("students.xlsx")

# State management for two-stage verification
waiting_for_face = False
scanned_id = None
verification_start_time = None
VERIFICATION_TIMEOUT = 10  # seconds to show face

while True:
    ret, frame = cap.read() # get frame object from read function as mentioned by opencv documentation. 'ret' is simply a boolean value which tells if it was successful or not ( true or false )
    
    # STAGE 1: Barcode Scanning Mode
    if not waiting_for_face:
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY) # convert to gray as it increases accuracy for barcode detection
        barcodes = decode(gray) # returns a list even if one barcode is scanned

        for barcode in barcodes:
            barcode_data = barcode.data.decode("utf-8") # decode the data into readable format. The standard is utf-8
            print("Scanned ID:", barcode_data)

            if barcode_data in valid_cms_ids: # Valid barcode - switch to face verification mode
                scanned_id = barcode_data
                
                # Visual feedback | From documentation to display a box around the barcode being scanned
                x, y, w, h = barcode.rect # get coordinates of barcode to display a box around it
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, "BARCODE OK - SHOW FACE", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

                waiting_for_face = True
                verification_start_time = time.time()
                print(f"Barcode accepted! Please show your face for verification...")
            else:
                # Invalid barcode
                x, y, w, h = barcode.rect
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 0, 255), 2)
                cv2.putText(frame, "ACCESS DENIED!!", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
    
    # STAGE 2: Face Verification Mode
    else:
        elapsed_time = time.time() - verification_start_time
        remaining_time = int(VERIFICATION_TIMEOUT - elapsed_time)
        
        # Check for timeout
        if elapsed_time > VERIFICATION_TIMEOUT:
            print("Verification timeout - please scan again")
            waiting_for_face = False
            scanned_id = None
            continue # skip rest of the loop & go back to top, if timeout
        
        # Display countdown
        cv2.putText(frame, f"Verifying face for ID: {scanned_id}", (10, 30), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
        cv2.putText(frame, f"Time remaining: {remaining_time}s", (10, 60), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 0), 2)
        
        # Try to detect and verify faced showing in live camera
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        live_faces = face_recognition.face_encodings(rgb_frame)
        
        if live_faces:
            live_encoding = live_faces[0]
            stored_encoding = face_database.get(scanned_id)

            if stored_encoding is not None:
                match = face_recognition.compare_faces([stored_encoding], live_encoding)[0]

                if match:
                    result = "ACCESS GRANTED"
                    color = (0, 255, 0)
                    print(result)
                    
                    # Display success message
                    cv2.putText(frame, result, (10, 120), 
                               cv2.FONT_HERSHEY_SIMPLEX, 1.2, color, 3)
                    cv2.imshow("Barcode Scanner", frame) # Refresh window after putting the text
                    cv2.waitKey(3000)  # Show success message for 3 seconds
                    
                    # Reset for next person
                    waiting_for_face = False
                    scanned_id = None
                else:
                    result = "FACE MISMATCH"
                    color = (0, 0, 255)
                    print(result)
                    cv2.putText(frame, result, (10, 120), 
                               cv2.FONT_HERSHEY_SIMPLEX, 1.2, color, 3)
                    cv2.imshow("Barcode Scanner", frame)
                    cv2.waitKey(3000)
                    
                    # Reset for next person
                    waiting_for_face = False
                    scanned_id = None
            else:
                result = "FACE NOT REGISTERED"
                color = (0, 0, 255)
                cv2.putText(frame, result, (10, 120), 
                           cv2.FONT_HERSHEY_SIMPLEX, 1.0, color, 2)
        else: # No face detected yet - keep waiting
            cv2.putText(frame, "Waiting for face...", (10, 120), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 0), 2)

    cv2.imshow("Barcode Scanner", frame) # Refresh window after putting the text

    if cv2.waitKey(1) == ord('q'): break # listen for input for 

# Safe way to exit the program as mentioned in OpenCV's documentation
cap.release()
cv2.destroyAllWindows()