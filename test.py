# What even is the Program??
# Problem is that I have to show my card to guard and lift my helmet to identify myself. I want to minimize the labor to save time for everyone.
# Solution is to implement a system that will scan the barcode of your card, and match it with your face. But I still have to lift my helmet tho.

# After the barcode is scanned, the system will ask for verification using face. If approved it will open the gate.

# It will display the students inside NUST, and you lift helmet once you go, so it will identify you and it will delete your entry from the students inside NUST

import cv2
from pyzbar.pyzbar import decode
import numpy as np
from openpyxl import load_workbook
import face_recognition
import os

def load_cms_ids(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active

    cms_ids = []

    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            cms_ids.append(str(row[0]).strip())

    return cms_ids

def load_face_database(folder_path):
    face_db = {}

    for filename in os.listdir(folder_path):
        if filename.lower().endswith((".jpg", ".png", ".jpeg")):
            cms_id = os.path.splitext(filename)[0]
            image_path = os.path.join(folder_path, filename)

            try:
                # Load with OpenCV first
                bgr_image = cv2.imread(image_path)

                if bgr_image is None:
                    print(f"❌ Could not read image: {filename}")
                    continue

                # Convert BGR → RGB (CRITICAL STEP)
                rgb_image = cv2.cvtColor(bgr_image, cv2.COLOR_BGR2RGB)
                

                encodings = face_recognition.face_encodings(rgb_image)

                if encodings:
                    face_db[cms_id] = encodings[0]
                    print(f"✅ Loaded face for CMS ID: {cms_id}")
                else:
                    print(f"⚠️ No face found in image: {filename}")

            except Exception as e:
                print(f"❌ Error processing {filename}: {e}")

    return face_db


face_database = load_face_database("faces")


cap = cv2.VideoCapture(0) # turns on webcam and uses default camera (0)

valid_cms_ids = load_cms_ids("./students.xlsx") # actual database
# valid_cms_ids = ["545895", "557955", "123456"] # temporary database


while True: # scan the camera until asked to shut down.
    ret, frame = cap.read() # using tuple unpacking. Python always returns tuples when there are multiple values being returned from a single function. ret is a boolean value while frame is the actual image data

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY) # convert the image to black and white since barcode's color don't matter. This increases detection accuracy.

    barcodes = decode(gray) # detect barcode in the given image. If found returns decoded objects, if not then empty list

    for barcode in barcodes: # if multiple barcodes appear, use loop
        barcode_data = barcode.data.decode("utf-8") # decodes the barcode data into human readable text
        print("Scanned ID:", barcode_data)

        if barcode_data in valid_cms_ids: # check if database id matches what is shown
            result = "ACCESS GRANTED"
            color = (0, 255, 0)

            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            live_faces = face_recognition.face_encodings(rgb_frame)

            print(live_faces)
            
            if live_faces:
                live_encoding = live_faces[0]
                stored_encoding = face_database.get(barcode_data)

                if stored_encoding is not None:
                    match = face_recognition.compare_faces(
                        [stored_encoding], live_encoding
                    )[0]

                    if match:
                        result = "ACCESS GRANTED"
                        color = (0, 255, 0)
                    else:
                        result = "FACE MISMATCH"
                        color = (0, 0, 255)
                else:
                    result = "FACE NOT REGISTERED"
                    color = (0, 0, 255)
            else:
                result = "NO FACE DETECTED"
                color = (0, 0, 255)

        else:
            result = "ACCESS DENIED"
            color = (0, 0, 255)

        # LIVE FEEDBACK SYSTEM:

        x, y, w, h = barcode.rect # get exact coordinates of barcode on screen

        cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2) # draws a box around the barcode

        cv2.putText(frame, result, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, color, 2) # display result, whether access allowed or not


    cv2.imshow("Barcode Scanner", frame) # Open live camera window

    if cv2.waitKey(1) & 0xFF == ord('q'): # Press q to shut down the app rather than kill terminal all the time
        break

cap.release()
cv2.destroyAllWindows()







